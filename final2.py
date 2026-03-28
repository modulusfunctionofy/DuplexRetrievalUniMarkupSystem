from sentence_transformers import SentenceTransformer, util
import PyPDF2
import os
from openpyxl import Workbook

def extract_text_from_pdf(pdf_path):
    text = ""
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    return text

def folder_reader(folder_path):
    st_names = []
    st_ans = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(folder_path, filename)
            st_names.append(filename)                
            st_ans.append(extract_text_from_pdf(pdf_path))  
    return st_names,st_ans

# model = SentenceTransformer('all-MiniLM-L6-v2') 
# model.save('./offline_model')

model = SentenceTransformer('./offline_model')

ans_real = extract_text_from_pdf('folderUSER/lfimgsUSER/ex1.pdf')


folder_path = 'folderUSER/rtimgsUSER'  #os.path.dirname(os.path.abspath(__file__))

name,ans = folder_reader(folder_path) 


emb_ans = []
emb_real = model.encode(ans_real, convert_to_tensor=True)


for i in ans:
    emb_ans.append(model.encode(i,convert_to_tensor=True))

result = []

for i in emb_ans:
    result.append(util.cos_sim(i, emb_real).item())


for i in range(len(result)):
    print(name[i], "=" ,result[i]*100)


wb = Workbook()
ws = wb.active

ws['A1'] = "Name"
ws['B1'] = "Score"

for i in range(len(result)):
    ws.append([name[i],result[i]])

wb.save(r"folderUSER/rtimgsUSER/example.xlsx")