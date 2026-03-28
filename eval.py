import nltk
import string
from sklearn.feature_extraction.text import TfidfVectorizer
from nltk.corpus import stopwords
import nltk.stem.porter
from sentence_transformers import SentenceTransformer, util
import PyPDF2
import os
from openpyxl import Workbook
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
import math


# links to update accordingly
answer_path = r'folderUSER/lfimgsUSER/ex1.pdf'                  # put answer path
folder_path = r'folderUSER/rtimgsUSER'                          # put folder path which contains answer of students
model_folder_path = './offline_model' # put path of folder of .py with folder
result_path = r'folderUSER/rtimgsUSER/result.xlsx'        # path for result


# model = SentenceTransformer('all-MiniLM-L6-v2') 
# model.save('model_folder_path')
ps = nltk.stem.porter.PorterStemmer()


def extract_text_from_pdf(pdf_path):
    text = ""
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    return text

def folder_reader(folder_path):
    st_names = []
    st_ans = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(folder_path, filename)
            st_names.append(filename)                
            st_ans.append(extract_text_from_pdf(pdf_path))  
    return st_names,st_ans

def toLower(text):
    if not isinstance(text, str):
        text = str(text)
    return text.lower()

def remove_punc(text):
    exclude = string.punctuation
    if not isinstance(text, str):
        text = str(text)
    return text.translate(str.maketrans('','',exclude))

def remove_stopword(text):
    new_text = []
    for word in text.split():
        if word in stopwords.words('english'):
            new_text.append('')
        else:
            new_text.append(word)
    x = new_text[:]
    new_text.clear()
    return " ".join(x)
              
def stemming(text):
    if not isinstance(text, str):
        text = str(text)
    return " ".join(ps.stem(word) for word in text.split())

def tokenize(text):                    
    return text.split()

def preprocessing(text):
    text = toLower(text)
    text = remove_punc(text)
    text = remove_stopword(text)
    text = stemming(text)
    text = tokenize(text)
    return text

def semantic_emb(ans):
    emb_ans = []
    for i in ans:
        emb_ans.append(model.encode(i,convert_to_tensor=True))
    return emb_ans

def cos_sim_score(emb_key,emb_ans):
    result = []
    for i in emb_ans:
        result.append(util.cos_sim(i, emb_key).item())
    return result

def batch_preprocessing(ans):
    ans_preprocessed = []
    for i in ans:
        ans_preprocessed.append(preprocessing(i))
    return ans_preprocessed

def tfidf_similarity(text, text_list):
    text = " ".join(text)
    text_list = [" ".join(t) for t in text_list]

    corpus = [text] + text_list
    vectors = vectorizer.fit_transform(corpus)

    input_vec = vectors[0]
    list_vecs = vectors[1:]

    similarities = cosine_similarity(input_vec, list_vecs).flatten()

    return similarities.tolist()

def give_marks(bert_score,tfidf_score,sentence_level_score):
    marks = []
    for i in range(len(bert_score)):
        marks.append(min(math.ceil((0.6*float(bert_score[i]) + 0.15*float(tfidf_score[i]) + 0.25*float(sentence_level_score[i]))*10),10))
    return marks
    
def sentence_sim_for_each(emb_key1,answer):
    answ = answer.split('.')
    emb_answ = semantic_emb(answ)
    s = []
    for i in emb_answ:
        p = []
        for j in emb_key1:
            p.append(util.cos_sim(i, j).item())
        s.append(max(p))
    return sum(s)/len(s)

def sentence_level_sim(key,ans):
    r = []
    key_s = key.split('.')
    emb_key2 = semantic_emb(key_s)
    for i in ans:
        r.append(sentence_sim_for_each(emb_key2,i))
    return r

# extraction of files
key = extract_text_from_pdf(answer_path)
name,ans = folder_reader(folder_path)


# model imported
vectorizer = TfidfVectorizer()
model = SentenceTransformer(model_folder_path)


# semantic analysis 
# embedding
emb_key = model.encode(key,convert_to_tensor=True)
emb_ans = semantic_emb(ans)
# similarity check
semantic_score = cos_sim_score(emb_key,emb_ans)


# sentnce level similarity
sentence_level_score = sentence_level_sim(key,ans)


#word similarity
#preprocessing
key_preprocessed = preprocessing(key)
ans_preprocessed = batch_preprocessing(ans)
# tfidf vector similarity
tfidf_similarity_score = tfidf_similarity(key_preprocessed,ans_preprocessed)


# marking
marks = give_marks(semantic_score,tfidf_similarity_score,sentence_level_score)


# printing
print(semantic_score)
print("-----------------------------------")
print(tfidf_similarity_score)
print("-----------------------------------")
print(sentence_level_score)
print("-----------------------------------")
print(marks)


# creating result file
wb = Workbook()
ws = wb.active
ws['A1'] = "Name"
ws['B1'] = "Score"
for i in range(len(marks)):
    ws.append([name[i],marks[i]])
wb.save(result_path)